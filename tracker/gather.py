"""Gather every company from every source into one record per domain.

Precedence for status is by RECENCY of the verdict, newest first: the 1 Sep
campaign files decide, then master v3, then the 31 Aug passes. A company DQ'd
on 31 Aug but re-listed on 1 Sep is active; the reverse is DQ.
"""
import re, json, csv
from openpyxl import load_workbook

U = '/root/.claude/uploads/cdc2f6cb-7804-5be0-bfc8-f894363d3735/'
R = '/home/user/Master-Target-List/'
F = {
 'master':  U + 'cee59e13-Hunter_Power_Master_Target_List_20260825_v3.xlsx',
 'newt':    U + 'd2024f6e-Hunter_Power_New_Targets_for_Campaigns_20260901.xlsx',
 'valid':   U + '6500def4-Hunter_Power_Validated_Campaigns_20260901.xlsx',
 'grata':   U + '48c5eb96-Hunter_Power_Grata_Expansion_20260831.xlsx',
 'campasn': U + 'a9cb6c9e-Hunter_Power_Campaign_Assignment_20260831.xlsx',
 'reass':   U + '7531be30-Hunter_Power_Grata_Expansion_Reassessment_20260831.xlsx',
 'ben':     U + '0c4c04d5-Leads_from_Ben__Anchor__with_Olsens_Campaign_Suggestion.csv',
}

def norm(d):
    if not d: return None
    d = str(d).strip().lower()
    d = re.sub(r'^https?://', '', d); d = re.sub(r'^www\.', '', d)
    d = d.split('/')[0].split('?')[0].strip()
    return d or None

def sheet(path, name, hr):
    wb = load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close(); return []
    ws = wb[name]
    hdr = [c.value for c in next(ws.iter_rows(min_row=hr, max_row=hr))]
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
    out = []
    for r in ws.iter_rows(min_row=hr + 1, values_only=True):
        if any(r): out.append({k: r[i] for k, i in idx.items()})
    wb.close(); return out

DB = {}
def put(dom, **kw):
    if not dom: return None
    rec = DB.setdefault(dom, {'domain': dom, 'sources': []})
    for k, v in kw.items():
        if k == 'source':
            if v and v not in rec['sources']: rec['sources'].append(v)
        elif v not in (None, '', '[unknown]') and not rec.get(k):
            rec[k] = v
    return rec

CAMPS_NT = ['Commercial & Industrial Ctrs','Motor & Generator Companies','Utility Line & Substation',
 'Mission critical power','Traffic Signals & Lighting','Testing & Commissioning',
 'Electrical Apparatus Servicing','EV Charging','Lightning Protection','Electrical Equipment Mfg',
 'Power Systems Engineering']

# ---- master v3 -------------------------------------------------------------
MSHEETS = [('Batch 1 (250)','Batch 1'),('Batch 2 (250)','Batch 2'),('Batch 3 (147)','Batch 3'),
           ('NEW - Add to batch 1 or 2 (11)','New add B1/2'),('NEW - Add to batch 2 or 3 (51)','New add B2/3'),
           ('Nurture (below floor)','Nurture'),('DQ log','DQ')]
for sh, lab in MSHEETS:
    for r in sheet(F['master'], sh, 2):
        put(norm(r.get('Website')), source=f'Master v3 / {lab}',
            company=r.get('Company'), hq=r.get('HQ'), band=r.get('Band'),
            bucket=r.get('Type bucket'), priority=r.get('Priority'), rank=r.get('Rank'),
            verification=r.get('Verification'), scope=r.get('Scope'), model=r.get('Model'),
            size=r.get('Size estimate and basis'), age=r.get('Age'), ownership=r.get('Ownership'),
            why=r.get('Why / verdict reason'), unknowns=r.get('Unknowns'),
            owner=r.get('Owner / principal'), angle=r.get('Outreach angle'),
            neta=r.get('NETA status'), acq=r.get('Acquisition / parent check'),
            confidence=r.get('Confidence'), contact=r.get('Contact'),
            contact_title=r.get('Contact title'), contact_email=r.get('Contact email'),
            contact_phone=r.get('Contact phone'), src_list=r.get('Source list'),
            description=r.get('Inven description (AI summary)'),
            master_status=lab)

# ---- 31 Aug reassessment (best descriptions) -------------------------------
for r in sheet(F['reass'], 'Full reassessment', 2):
    put(norm(r.get('Domain')), source='Reassessment 31 Aug', company=r.get('Company'),
        hq=r.get('HQ'), revenue=r.get('Revenue est.'), employees=r.get('Employee est.'),
        ownership=r.get('Ownership'), founded=r.get('Year founded'),
        bucket_reassessed=r.get('Reassessed bucket'), priority_reassessed=r.get('Reassessed priority'),
        reass_action=r.get('Action'), reass_reason=r.get('Reassessment reasons'),
        description=r.get('Description'), wave=r.get('Wave'), profile=r.get('Profile URL'))

# ---- 31 Aug campaign assignment -------------------------------------------
for sh in ['Testing & Commissioning','Electrical Apparatus Servicing','Utility Line & Substation',
           'Commercial & Industrial Ctrs','Mission critical power','Motor & Generator Companies',
           'Traffic Signals & Lighting','EV Charging','Lightning Protection',
           'Electrical Equipment Mfg','Power Systems Engineering','Nurture','DQ']:
    for r in sheet(F['campasn'], sh, 4):
        put(norm(r.get('Domain') or r.get('Website')), source=f'Campaign Assignment 31 Aug / {sh}',
            company=r.get('Company'), hq=r.get('HQ'), revenue=r.get('Revenue est.'),
            employees=r.get('Employee est.'), ownership=r.get('Ownership'),
            founded=r.get('Year founded'), description=r.get('Description'),
            ca_sheet=sh, ca_reason=r.get('Reason') or r.get('Reasons') or r.get('Note'))

# ---- 1 Sep validated + new targets ----------------------------------------
for key, lab in [('valid','Validated 1 Sep'), ('newt','New Targets 1 Sep')]:
    for sh in CAMPS_NT + (['NEW additions (138)'] if key == 'newt' else []) + ['Remove - DQ']:
        for r in sheet(F[key], sh, 2):
            put(norm(r.get('Domain')), source=f'{lab} / {sh}',
                company=r.get('Company'), hq=r.get('HQ'), bucket_v=r.get('Bucket'),
                v_campaign=(None if sh == 'Remove - DQ' else (r.get('Correct campaign') or sh)),
                v_sheet=sh, v_confidence=r.get('confidence'), v_reason=r.get('reason'),
                size_read=r.get('Size read'), size_check=r.get('Size check'),
                web_check=r.get('Web check'), contact=r.get('Contact'),
                contact_title=r.get('Contact title'), contact_email=r.get('Contact email'),
                angle=r.get('Outreach angle'), priority_v=r.get('Priority'),
                employees=r.get('Employee est.'), revenue=r.get('Revenue est.'),
                founded=r.get('Year founded'), ownership=r.get('Ownership'),
                description=r.get('Description'), profile=r.get('Profile URL'),
                dq_reason=(r.get('reason') if sh == 'Remove - DQ' else None))

# ---- Ben / Anchor leads ----------------------------------------------------
for r in csv.DictReader(open(F['ben'])):
    em = (r.get('Email') or '').strip()
    d = norm(em.split('@')[-1]) if '@' in em else None
    nm = ' '.join(x for x in [(r.get('First Name') or '').strip(), (r.get('Last Name') or '').strip()] if x)
    put(d, source='Ben/Anchor leads', company=(r.get('Company') or '').strip(),
        hq=(r.get('Location') or '').strip(), contact=nm or None,
        contact_title=(r.get('Title') or '').strip() or None, contact_email=em or None,
        contact_phone=(r.get('Phone') or '').strip() or None,
        ben_tier=(r.get('Tier') or '').strip(), ben_reply=(r.get('Prior reply status') or '').strip(),
        ben_succession=(r.get('Succession Score') or '').strip(),
        why=(r.get('Why it fits') or '').strip(), ben_industry=(r.get('Source industry') or '').strip(),
        ben_campaign=(r.get('Campaign') or '').strip())

# ---- association free lists (raw leads) ------------------------------------
try:
    AN = json.load(open(R + 'scripts/xref/association_netnew.json'))
    for r in AN['net_new']:
        put(norm(r.get('domain')), source='Trade association directory',
            company=r.get('name'), hq=r.get('headquarters'), employees=r.get('employees'),
            ownership=r.get('ownership'), founded=r.get('year_founded'),
            description=r.get('description'), assoc_list=r.get('source_list'))
except FileNotFoundError:
    pass

json.dump(DB, open('db.json', 'w'), indent=1, default=str)
print('unique companies gathered:', len(DB))
import collections
print('rows carrying a description:', sum(1 for v in DB.values() if v.get('description')))
print('top sources:', collections.Counter(s.split(' / ')[0] for v in DB.values() for s in v['sources']).most_common())
