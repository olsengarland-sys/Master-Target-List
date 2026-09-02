import json, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(open('rows.json')); rows = D['rows']
HDRF = PatternFill('solid', fgColor='1F3864'); HF = Font(color='FFFFFF', bold=True, size=10)
GREEN = PatternFill('solid', fgColor='E2EFDA'); AMBER = PatternFill('solid', fgColor='FFF2CC')
RED = PatternFill('solid', fgColor='FCE4E4'); GREY = PatternFill('solid', fgColor='F2F2F2')
WRAP = Alignment(wrap_text=True, vertical='top'); BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)

COLS = [
 ('Status',22,'status'), ('Decided by',26,'decided_by'), ('Campaign',26,'campaign'), ('Priority',12,'priority'),
 ('Company',32,'company'), ('Domain',26,'domain'), ('HQ',24,'hq'),
 ('Postal address',40,'address'), ('Address usable?',13,'address_mailable'),
 ('Bucket',22,'bucket'), ('Band',14,'band'), ('Rank',8,'rank'),
 ('Owner / contact',20,'owner_contact'), ('Owner title',26,'owner_title'),
 ('Email',30,'email'), ('Email verified',12,'email_verified'),
 ('MOBILE',18,'mobile'), ('Direct dial',18,'direct_dial'), ('Office phone',18,'office_phone'),
 ('Other contacts',24,'other_contacts'),
 ('Revenue est.',26,'revenue'), ('Employees',11,'employees'), ('Year founded',11,'founded'),
 ('Ownership',16,'ownership'), ('Size read',30,'size_read'), ('Size check',30,'size_check'),
 ('Scope',12,'scope'), ('Model',12,'model'), ('Verification',22,'verification'),
 ('Web check',24,'web_check'), ('Confidence',12,'confidence'),
 ('Why / verdict reason',50,'why'), ('DQ reason',44,'dq_reason'),
 ('Reassessment action',22,'reass_action'), ('Reassessment reason',40,'reass_reason'),
 ('NETA status',14,'neta'), ('Acquisition / parent check',26,'acq_check'),
 ('Unknowns',26,'unknowns'), ('Outreach angle',40,'angle'),
 ('Ben tier',20,'ben_tier'), ('Ben reply status',14,'ben_reply'), ('Succession score',12,'ben_succession'),
 ('Conferences attended',44,'conferences'), ('QC FLAG - check before contact',40,'qc_flag'),
 ('Discovery wave',22,'wave'), ('Association list',34,'assoc_list'),
 ('Grata/Inven profile',30,'profile'), ('Source files (full trail)',70,'sources'),
 ('Description',80,'description_1'),
]
for i in range(2, D['maxparts'] + 1):
    COLS.append((f'Description (cont. {i})', 80, f'description_{i}'))

def table(ws, data, start=1):
    for i, (h, w, _) in enumerate(COLS, 1):
        c = ws.cell(row=start, column=i, value=h); c.fill = HDRF; c.font = HF; c.alignment = WRAP
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=6)
    ws.auto_filter.ref = f'A{start}:{get_column_letter(len(COLS))}{start + len(data)}'
    for n, r in enumerate(data, start + 1):
        for i, (_, _, k) in enumerate(COLS, 1):
            ws.cell(row=n, column=i, value=r.get(k)).alignment = WRAP
        st = r['status']
        fill = RED if st == 'DQ' else (GREY if st.startswith('RAW') else (AMBER if st == 'NURTURE' else None))
        if fill: ws.cell(row=n, column=1).fill = fill
        if r.get('mobile'): ws.cell(row=n, column=17).fill = GREEN
        if r.get('qc_flag'): ws.cell(row=n, column=44).fill = RED
        if str(r.get('ben_reply') or '').lower() == 'replied':
            for i in range(1, 5): ws.cell(row=n, column=i).fill = GREEN
        ws.row_dimensions[n].height = 26

order = {'ACTIVE TARGET': 0, 'NURTURE': 1, 'RAW LEAD - not screened': 2, 'DQ': 3, 'UNCLASSIFIED': 4}
rows.sort(key=lambda r: (order.get(r['status'], 9), not r.get('mobile'), not r.get('email'),
                         str(r.get('priority') or 'zz'), str(r.get('company') or '')))
live = [r for r in rows if r['status'] != 'DQ']
CAMPAIGNS = ['Commercial & Industrial Ctrs','Motor & Generator Companies','Utility Line & Substation',
 'Mission critical power','Traffic Signals & Lighting','Testing & Commissioning',
 'Electrical Apparatus Servicing','EV Charging','Lightning Protection','Electrical Equipment Mfg',
 'Power Systems Engineering']

wb = Workbook(); ws = wb.active; ws.title = '1. Summary'
ws.cell(row=1, column=1, value='Hunter Power Services - Master Tracker').font = TITLE
ws.cell(row=2, column=1, value='Single source of truth. Every company found across every source, with every '
        'detail held for it. Generated 2 September 2026.').alignment = WRAP
n_st = collections.Counter(r['status'] for r in rows)
n_cp = collections.Counter(r['campaign'] or '(none)' for r in live)
r = 4
def S(cells, bold=False, fill=None):
    global r
    for i, v in enumerate(cells, 1):
        c = ws.cell(row=r, column=i, value=v); c.alignment = WRAP
        if bold: c.font = BOLD
        if fill: c.fill = fill
    r += 1
S(['THE UNIVERSE', 'Companies', 'Note'], True, HDRF)
for i in range(1, 4): ws.cell(row=r-1, column=i).font = HF
S(['Total companies tracked', len(rows), 'every unique domain across all seven sources'], True)
S(['ACTIVE TARGET', n_st.get('ACTIVE TARGET', 0), 'in a campaign, any priority - work these'], False, GREEN)
S(['NURTURE', n_st.get('NURTURE', 0), 'right model, below the size floor'], False, AMBER)
S(['RAW LEAD - not screened', n_st.get('RAW LEAD - not screened', 0),
   'from free trade-association directories; never screened against the thesis'], False, GREY)
S(['DQ', n_st.get('DQ', 0), 'a gate fired somewhere in the chain - logged so none is re-screened cold'], False, RED)
r += 1
S(['CONTACTABILITY (non-DQ only)', 'Companies', 'Note'], True, HDRF)
for i in range(1, 4): ws.cell(row=r-1, column=i).font = HF
S(['Non-DQ companies', len(live), ''], True)
S(['with an owner MOBILE', sum(1 for x in live if x.get('mobile')), 'the highest-value field'])
S(['with an email', sum(1 for x in live if x.get('email')), ''])
S(['with a postal address', sum(1 for x in live if x.get('address')), ''])
S(['carrying a QC flag', sum(1 for x in live if x.get('qc_flag')), 'check before contact'])
S(['seen at a conference', sum(1 for x in live if x.get('conferences')), 'warm-intro route'])
S(['REPLIED to prior outreach', sum(1 for x in live if str(x.get('ben_reply') or '').lower() == 'replied'),
   'live conversations - highest priority'], True, GREEN)
r += 1
S(['BY CAMPAIGN (non-DQ, all priorities)', 'Companies', 'Tab'], True, HDRF)
for i in range(1, 4): ws.cell(row=r-1, column=i).font = HF
for i, c in enumerate(CAMPAIGNS, 3):
    S([c, n_cp.get(c, 0), f'{i}. {c[:24]}'])
S(['(no campaign assigned)', n_cp.get('(none)', 0), 'Unscreened leads tab'], False, GREY)
r += 1
S(['SOURCES FOLDED IN', '', ''], True, HDRF)
for i in range(1, 4): ws.cell(row=r-1, column=i).font = HF
for nm, note in [
  ('Master Target List v3 (25 Aug)', 'the base: 1,922 in the master tab plus 62 on the two "NEW - Add to batch" sheets'),
  ('Grata Expansion (31 Aug)', '1,755 raw candidates - a search result, not a verdict'),
  ('Reassessment (31 Aug)', 'every candidate re-bucketed and re-prioritised; best descriptions come from here'),
  ('Campaign Assignment (31 Aug)', 'adjudicated all 1,755: 832 campaign + 212 Nurture + 711 DQ'),
  ('Validated Campaigns (1 Sep)', 'validated 1,045 rows: kept 729, DQd 316'),
  ('New Targets for Campaigns (1 Sep)', 'the 729 plus 138 released from master Nurture = 867'),
  ('Ben / Anchor leads', '47 named contacts, 2 of whom have already replied'),
  ('Trade association directories', '1,033 free net-new names - unscreened'),
  ('Contact enrichment (Inven + Grata)', 'owner mobiles, emails, postal addresses, QC flags'),
  ('Conference rosters', '32 events; flags which targets are in the room')]:
    S([nm, '', note])
r += 1
S(['HOW TO READ IT', '', ''], True, HDRF)
for i in range(1, 4): ws.cell(row=r-1, column=i).font = HF
for t in ['Status is decided by the NEWEST verdict. A company DQd on 31 August but re-listed on 1 September is active; the reverse is DQ.',
          'The "Decided by" column on the tracker names the file that set each status, and "Source files" carries the full trail.',
          'Campaign tabs hold every non-DQ company in that campaign at ANY priority, with the same columns as the tracker.',
          'Green = an owner mobile is held. Red = DQ, or a QC flag. Amber = Nurture. Grey = an unscreened raw lead.',
          'RAW LEADs have not been screened against the thesis and will contain non-targets - screen before working them.']:
    S([t, '', ''])
ws.column_dimensions['A'].width = 46; ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 96

ws = wb.create_sheet('2. MASTER TRACKER'); table(ws, rows)
for i, c in enumerate(CAMPAIGNS, 3):
    sub = [x for x in live if x['campaign'] == c]
    w = wb.create_sheet(f'{i}. {c}'[:31])
    w.cell(row=1, column=1, value=f'{c} - {len(sub)} companies, not DQd, all priorities').font = TITLE
    table(w, sub, start=3)
sub = [x for x in live if not x['campaign']]
w = wb.create_sheet('14. Unscreened leads')
w.cell(row=1, column=1, value=f'{len(sub)} leads with no campaign - mostly free trade-association '
       'names never screened against the thesis. Screen before working.').font = TITLE
table(w, sub, start=3)

OUT = '/home/user/Master-Target-List/Hunter_Power_MASTER_TRACKER_20260902.xlsx'
wb.save(OUT)
print('tabs:', len(wb.sheetnames)); print('wrote', OUT)
