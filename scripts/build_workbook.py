import json, re, statistics, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CAND = json.load(open('scripts/candidates.json'))
COMPS = json.load(open('scripts/comps.json'))
CUTS  = json.load(open('scripts/cuts.json'))

OUT = 'Hunter_Power_Grata_Expansion_20260831.xlsx'
ORDER = ['T1','T2','T3','T4','T5','T6','S7','S8','S9','S10']
NAMES = {
 'T1':'T1 Testing & Commissioning','T2':'T2 Apparatus Service','T3':'T3 Utility Line & Substation',
 'T4':'T4 Industrial Plant Electrical','T5':'T5 Commercial & Institutional','T6':'T6 Critical Power & Data Ctr',
 'S7':'S7 Motor & Generator','S8':'S8 Traffic Signals & Lighting','S9':'S9 EV Charging',
 'S10':'S10 Lightning Protection'}

RED   = PatternFill('solid', fgColor='FCE4E4')
AMBER = PatternFill('solid', fgColor='FFF2CC')
HDR   = PatternFill('solid', fgColor='1F3864')
HDRF  = Font(color='FFFFFF', bold=True, size=10)
BOLD  = Font(bold=True)
TITLE = Font(bold=True, size=13)
WRAP  = Alignment(wrap_text=True, vertical='top')
TOP   = Alignment(vertical='top')
THIN  = Border(*[Side(style='thin', color='D0D0D0')]*4)

wb = Workbook()

def header(ws, row, cols):
    for i,c in enumerate(cols,1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = WRAP
    ws.freeze_panes = ws.cell(row=row+1, column=1)

def widths(ws, w):
    for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width = x

def money(v):
    if v is None: return '[unknown]'
    return '$%.1fM' % (v/1_000_000)

def num(v):
    return '[unknown]' if v is None else v

# ---------------------------------------------------------------- Summary
ws = wb.active; ws.title = 'Summary'
ws['A1'] = 'Hunter Power Services — Grata Bucket Expansion'; ws['A1'].font = Font(bold=True, size=15)
ws['A2'] = 'Run date 2026-08-31 · Sourcing analyst pass · Parts A (expansion), B (public comps), C (market cuts)'
ws['A3'] = 'All company facts below are Grata returned fields or the master list. Revenue and employee figures are GRATA ESTIMATES. Unknowns are written [unknown], never guessed.'
ws['A3'].font = Font(italic=True, size=9)

r = 5
ws.cell(row=r, column=1, value='NEW CANDIDATES BY BUCKET AND PROVISIONAL PRIORITY').font = TITLE; r += 1
cols = ['Bucket','Bucket name','P2-candidate','P3-candidate','Nurture-candidate','DQ-candidate','Total new','Dropped as known (dedupe)','Dropped off-thesis at triage']
header(ws, r, cols); r += 1
tot = collections.Counter()
for b in ORDER:
    v = CAND[b]; pc = collections.Counter(c['priority'] for c in v['candidates'])
    n = len(v['candidates'])
    row = [b, NAMES[b], pc['P2-candidate'], pc['P3-candidate'], pc['Nurture-candidate'], pc['DQ-candidate'],
           n, v.get('dropped_known', 0), v.get('dropped_offthesis', 0)]
    for i,x in enumerate(row,1): ws.cell(row=r, column=i, value=x).alignment = TOP
    for k in ('P2-candidate','P3-candidate','Nurture-candidate','DQ-candidate'): tot[k] += pc[k]
    tot['n'] += n; tot['dk'] += v.get('dropped_known',0); tot['do'] += v.get('dropped_offthesis',0)
    r += 1
trow = ['TOTAL','', tot['P2-candidate'], tot['P3-candidate'], tot['Nurture-candidate'], tot['DQ-candidate'], tot['n'], tot['dk'], tot['do']]
for i,x in enumerate(trow,1):
    c = ws.cell(row=r, column=i, value=x); c.font = BOLD
r += 2
W2 = json.load(open('scripts/xref/wave2.json'))
ws.cell(row=r, column=1, value='WAVE 2 - INVEN x GRATA CROSS-REFERENCE (lower precision by design; see gaps)').font = TITLE; r += 1
header(ws, r, ['Bucket','Bucket name','P2-candidate','P3-candidate','Nurture-candidate','DQ-candidate','Total new','Dropped as known','Dropped off-thesis by verifier'])
r += 1
w2t = collections.Counter()
for b in ORDER:
    t = W2['tally'][b]
    row = [b, NAMES[b], t['P2-candidate'], t['P3-candidate'], t['Nurture-candidate'], t['DQ-candidate'], t['new'], t['dropped_known'], t['dropped_offthesis']]
    for i,x in enumerate(row,1): ws.cell(row=r, column=i, value=x).alignment = TOP
    for k in ('P2-candidate','P3-candidate','Nurture-candidate','DQ-candidate'): w2t[k] += t[k]
    w2t['n'] += t['new']; w2t['dk'] += t['dropped_known']; w2t['do'] += t['dropped_offthesis']
    r += 1
for i,x in enumerate(['TOTAL','', w2t['P2-candidate'], w2t['P3-candidate'], w2t['Nurture-candidate'], w2t['DQ-candidate'], w2t['n'], w2t['dk'], w2t['do']],1):
    ws.cell(row=r, column=i, value=x).font = BOLD
r += 1
c = ws.cell(row=r, column=1, value=('Wave 2 method: per bucket, one Inven semantic search (top 200 relevance-ranked rows, the three saved Inven lists excluded engine-side) '
    'plus two fresh Grata keyword angles wave 1 never ran; deterministic dedupe against master, DQ log, Nurture, screened log, outreach list AND wave-1 candidates; '
    'then an adversarial verifier per bucket challenged every priority. %d cross-bucket duplicates removed; %d non-US rows dropped in post-hoc cleanup. '
    "Wave-2 'dropped as known' also counts within-wave cross-source duplicates, so it overstates prior-list overlap.") % (W2['crossDupes'], W2['_cleanup']['foreign_count']))
c.alignment = WRAP; c.font = Font(italic=True, size=9)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); ws.row_dimensions[r].height = 45
r += 2
W3 = json.load(open('scripts/xref/wave3_recovery.json'))
ws.cell(row=r, column=1, value='WAVE 3 - REVENUE-ONLY RECOVERY (employee proxy removed per client rule)').font = TITLE; r += 1
header(ws, r, ['Bucket','Bucket name','P2-candidate','P3-candidate','DQ-candidate','Total new','Revenue-only universe','In-box w/ employee filter (wave 1)','Hidden by proxy'])
r += 1
w3t = collections.Counter()
for b in ORDER:
    v3 = W3['perBucket'][b]
    pc = collections.Counter(c['priority'] for c in v3['survivors'])
    inbox = CUTS[b]['in_box_5_50M']
    row = [b, NAMES[b], pc['P2-candidate'], pc['P3-candidate'], pc['DQ-candidate'], len(v3['survivors']),
           v3.get('count_revenue_only'), inbox, (v3.get('count_revenue_only') or 0) - inbox]
    for i,x in enumerate(row,1): ws.cell(row=r, column=i, value=x).alignment = TOP
    for k in ('P2-candidate','P3-candidate','DQ-candidate'): w3t[k] += pc[k]
    w3t['n'] += len(v3['survivors']); w3t['delta'] += (v3.get('count_revenue_only') or 0) - inbox
    r += 1
for i,x in enumerate(['TOTAL','', w3t['P2-candidate'], w3t['P3-candidate'], w3t['DQ-candidate'], w3t['n'], '', '', w3t['delta']],1):
    ws.cell(row=r, column=i, value=x).font = BOLD
r += 1
c = ws.cell(row=r, column=1, value=('Wave 3 re-ran each bucket\'s wave-1 primary keyword search with revenue $5-50M and NO employee filter, per the client rule that '
    'headcount is not an investment criterion (it was only a proxy). The proxy had been hiding ~%d in-box companies across buckets, concentrated in T6, T4 and S7. '
    'Only the top 1-2 pages per bucket were pulled; the T6/T4 tails remain unexplored.') % w3t['delta'])
c.alignment = WRAP; c.font = Font(italic=True, size=9)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); ws.row_dimensions[r].height = 40
r += 3

ws.cell(row=r, column=1, value='GRATA TOKEN BALANCE').font = TITLE; r += 1
for lab,val in [('Balance at start of run','5,743 of 7,000 remaining'),
                ('Balance after wave 1','4,758 of 7,000 remaining'),
                ('Balance after wave 2 (Inven x Grata cross-reference)','4,650 of 7,000 remaining'),
                ('Balance after wave 3 (revenue-only recovery)','4,614 of 7,000 remaining'),
                ('Consumed','985 (wave 1) + 72 (wave 2) + 36 (wave 3) Grata tokens'),
                ('Billing period','2026-08-24 to 2027-08-23'),
                ('Tools used','Grata: search_companies (3), find_similar_companies (3), get_public_comps (1), get_token_usage (0). Inven (wave 2): build_company_search, run_company_search, build_columns, get_lists - 31 visible screening credits total (1,452 -> 1,421) plus ~2,000 rows against the org export-volume quota (pool not visible via API). No contact, enrichment, CRM, list-write or export tool was called on either platform.')]:
    ws.cell(row=r, column=1, value=lab).font = BOLD
    ws.cell(row=r, column=2, value=val).alignment = WRAP
    r += 1
r += 2

ws.cell(row=r, column=1, value='CONFIDENCE AND GAPS').font = TITLE; r += 1
gaps = [
 ('Seller Intent unavailable',
  'seller_intent_levels: ["High Intent"] returns HTTP 403 — "Your organization does not have access to Seller Intent." Skipped per prompt instruction. No seller-intent cut exists in Part C for any bucket.'),
 ('hold_period_signal unavailable',
  'exit_readiness_signals: ["hold_period_signal"] is silently remapped by the API to transaction_signal (visible in the returned filters_used) and returns 0. Only founder_age_signal is usable. Treat the exit-readiness cut as founder-age only.'),
 ('Nurture-candidate is structurally empty',
  'The common filter employees_min: 10 removes sub-10-FTE firms before they can be returned, so the Nurture-candidate class cannot populate from these searches. Finding sub-floor nurture names requires a separate run with employees_min lowered — recommended as a follow-on.'),
 ('P1 was never assigned, by design',
  'Per the system doc §3.6, P1 requires the enrichment pass. Nothing here is above P2-candidate. Every priority in this workbook is provisional and read off a Grata description only.'),
 ('Seeds resolved by domain, not lookup_companies',
  'find_similar_companies and get_public_comps accept a domain directly as company_uid, so seeds were passed as domains. This saved ~300 tokens and 10 round trips but means no lookup_companies confirmation step ran on seed identity. Seed domains came from the master and are listed on the Public comps tab.'),
 ('Public comps map to listed PARENTS, not trade peers',
  'Grata resolves a private comp to whatever listed entity owns it. That produced non-comparable "comps" — Goldman Sachs (T6), Vistra (T6), KKR and HEICO (S7), ACCO Brands (S10), Ares and OFS Capital (S9). These are flagged amber on the Public comps tab and should be ignored for valuation.'),
 ('Scale discount is not optional',
  'Every multiple on the Public comps tab is an EMCOR/Quanta/Comfort Systems-scale listed multiple (13x-32x EV/EBITDA). These DO NOT transfer to a $2-10M EBITDA private target. Lower-middle-market electrical services trade in a materially lower band; the listed set is directional context for Olsen\'s pricing instinct, not a pricing model. The S7 seed 2 set (median 9.4x, genuine motor-service peers) is the closest thing in this run to a trade-honest read.'),
 ('locations filter deviates from the prompt literal',
  'locations: ["US"] returns HTTP 400 "Ambiguous locations term(s)". Country was passed as "United States" and states as "State, US" (e.g. "Texas, US"), which matches the prompt instruction to spell out state names.'),
 ('S9 EV Charging search is heavily off-thesis',
  'The EV-charging keyword profile returns vehicle dealerships, plumbing/HVAC and tree-service firms. 14 of 19 S9 survivors gate as DQ-candidate — consistent with the master\'s own S9 record (11 DQ of 17). S9 is not a productive hunting ground on Grata keywords alone.'),
 ('T2 bleeds into S7',
  '"electrical apparatus service" pulls EASA motor-shop results, overlapping S7. The two buckets should be read together, not as independent pipelines.'),
 ('Residential gate is literal, as the prompt specifies',
  'Any residential language in a Grata description fires the gate, including firms that list residential alongside commercial and industrial. 32 of 64 DQ-candidates fired on this gate. Several are likely C&I-dominant and worth a human re-read before they enter the DQ log permanently — the quoted phrase is on the DQ tab so the mix can be judged.'),
 ('Unknown-revenue pool is effectively zero',
  'In every bucket the five revenue bands sum to the unfiltered bucket total (variance 0-2 companies), i.e. Grata carries a revenue estimate for essentially every company in these filter sets. The pool was derived arithmetically rather than by 50 extra revenue_include_unknown calls.'),
 ('Outreach contacts list was added to dedupe AFTER the initial run',
  'The outreach contacts CSV (408 unique companies already contacted) was not among the inputs when the searches ran; it was supplied '
  'afterwards and folded into the dedupe index on a second pass. 27 candidates already on that outreach list were removed and are counted '
  'in each bucket\'s dropped-as-known figure. Contact names, titles and email addresses were used for matching only and appear nowhere in '
  'this workbook. Note also that 177 of the 408 outreach companies are in neither the master nor the screened log, so that list is itself '
  'an untracked source of names worth reconciling into the master.'),
 ('Search depth was not uniform across buckets',
  'Every bucket got a keyword search_companies pass AND a find_similar_companies pass seeded with up to 5 master P1/P2 companies '
  '(Grata warns per-seed signal dilutes past ~5 seeds). Page depth varied by how fast results went off-thesis: T2 and T3 ran 2 pages '
  'of similar results, the other buckets 1 page (25 rows). The prompt allows up to 4 pages; depth was cut where overlap with the '
  'keyword pass was already running above 40 percent, which it was in every bucket.'),
 ('S9 had no P1/P2 seed in the master',
  'S9 EV Charging holds no P1 or P2 company in the master, so the similar-to pass used the top P3 (Resound Energy, resoundenergy.com) '
  'as a fallback seed and the comps seed is the same P3. Treat S9 similarity results as anchored to a weaker seed than every other bucket.'),
 ('13 S9 results dropped off-thesis rather than logged as DQ',
  'The S9 similar-to pass returned pure LED-lighting-retrofit and energy-services firms with no EV or electrical-infrastructure service '
  'line. These were dropped at triage rather than carried into the DQ tab as noise; the names are recorded in scripts/candidates.json '
  'under S9.offthesis_note. No other bucket needed an off-thesis drop.'),
 ('WAVE 3: employee filter removed - proxy had hidden ~236 in-box companies',
  'Per the client rule that headcount is not a criterion, wave 3 re-ran the primary searches on revenue only. 87 new rows surfaced (33 P2), best names at headcounts the proxy excluded (L&S Electric at 360 FTE, sub-10-FTE critical-power specialists). The T6 (+110) and T4 (+46) revenue-only deltas were only sampled at the top of relevance ranking - paging deeper there is the cheapest remaining recovery. Caveat: with no headcount floor, some wave-3 rows may be estimate artifacts; revenue/employee plausibility was not re-checked.'),
 ('WAVE 2: treat as a broader, lower-precision pool',
  'Wave 2 (Inven semantic search + fresh Grata angles) trades precision for reach. Every priority was adversarially verified per bucket, but the completeness critic still recommends a human read of P2s before outreach. The wave-2 DQ pool (~544) contains verified gate misfires and must NOT be ingested into the DQ log wholesale.'),
 ('WAVE 2: T3 and T2 counts are inflated by drift',
  "T3 shows 137 new rows against a Grata-measured universe of ~102 - most of the excess is telecom/pipeline/heavy-civil drift; treat T3 wave-2 rows as unvetted. T2's 78 P2s against a ~273-firm universe is not credible either: an Inven refinement loosened the search ~10x (estimate 1,852 -> 19,739) and ~10 servo/PLC-repair names were kept on an unconfirmed assumption."),
 ('WAVE 2: T5 under-returned',
  'T5 delivered 19 P2 against a measured in-box universe of ~842. The Inven prompt drifted to MRO/distributor language. T5 needs a re-run with service-department-specific phrasing - the biggest single re-run opportunity.'),
 ('WAVE 2: only the top 200 Inven rows per bucket were examined',
  'Inven universes run 3,077-24,414 per bucket; only the top-200 relevance-ranked rows were pulled (2,000 export rows total). The unexamined tail plus never-exercised accreditation directories (NETA for T1, EASA/PEARL for T2/S7, LPI/UL Master Label for S10) are the highest-value next modalities.'),
 ('WAVE 2: ownership filtering is not comparable across buckets',
  'Inven ownership refinements resolved "not PE-backed" differently per bucket (VC-backed privates included in some, excluded in others - T1 and S10 excluded them). Per-bucket P2 pools are not comparable on ownership; Investor Backed rows carry REVIEW flags.'),
 ('WAVE 2: revenue estimates disagree across platforms by up to 4.9x',
  'Same-company revenue estimates from Inven vs Grata differ by up to 4.9x and cross the $5-50M box boundary (e.g. emoryelectric.com $40.9M vs $51.3M; krautomationinc.com $5.5M vs $27.1M). Box-based priority calls rest on soft estimates; verify revenue independently before acting on any single name.'),
 ('WAVE 2: known data defects cleaned or flagged',
  '12 non-US companies (foreign TLDs or critic-named) were dropped post-hoc; 7 rows carry REVIEW notes for critic-verified data errors (wrong description, implausible revenue/employee ratios, possible duplicate entity jhe-la.com/jbhcontractor.com); one invalid founding year (1694) was set to [unknown]. A Grata age-filter translation in T4 silently floored company age at 61 years, cutting pre-1965 family firms from the Grata-variant contribution.'),
 ('Description-only triage',
  'No enrichment call was made on any candidate. Model signals (recurring service revenue, contract base, technician count, NETA accreditation) are inferred from marketing prose. Expect meaningful false-positive and false-negative rates on the P2/P3 split.'),
]
header(ws, r, ['Gap / caveat','Detail']); r += 1
for g,d in gaps:
    ws.cell(row=r, column=1, value=g).alignment = WRAP
    c = ws.cell(row=r, column=2, value=d); c.alignment = WRAP; c.fill = AMBER
    ws.row_dimensions[r].height = 42
    r += 1
widths(ws, [34, 110, 16, 18, 20, 16, 12, 24, 26])

# ---------------------------------------------------------------- bucket tabs
BCOLS = ['Company','Domain','HQ (city, state)','Grata revenue est.','Grata employee est.','Ownership',
         'Year founded','Grata description (first 300 chars)','Source (search keywords or similar-to seed)',
         'Provisional priority','Gate / flag notes','Grata profile URL']
PRI_ORDER = {'P2-candidate':0,'P3-candidate':1,'Nurture-candidate':2,'DQ-candidate':3}

for b in ORDER:
    ws = wb.create_sheet(b)
    ws['A1'] = NAMES[b] + ' — new candidates not in master, DQ log, Nurture or screened log'
    ws['A1'].font = TITLE
    v = CAND[b]
    ws['A2'] = ('%d new candidates · %d dropped as already-known at dedupe · %d dropped off-thesis at triage · '
                'revenue and employee figures are Grata estimates'
                % (len(v['candidates']), v.get('dropped_known',0), v.get('dropped_offthesis',0)))
    ws['A2'].font = Font(italic=True, size=9)
    header(ws, 4, BCOLS)
    r = 5
    for c in sorted(v['candidates'], key=lambda x: (PRI_ORDER.get(x['priority'],9), -(x.get('rev') or 0))):
        desc = (c.get('desc') or '')[:300]
        row = [c['name'], c.get('domain') or '[unknown]', c.get('hq') or '[unknown]',
               money(c.get('rev')), num(c.get('emp')), c.get('own') or '[unknown]',
               num(c.get('yr')), desc, c.get('source',''), c['priority'], c.get('note',''), c.get('url','')]
        for i,x in enumerate(row,1):
            cell = ws.cell(row=r, column=i, value=x); cell.alignment = WRAP; cell.border = THIN
        if c['priority'] == 'DQ-candidate':
            for i in range(1, len(BCOLS)+1): ws.cell(row=r, column=i).fill = RED
        elif 'REVIEW' in c.get('note','') or c['priority'] == 'Nurture-candidate':
            for i in range(1, len(BCOLS)+1): ws.cell(row=r, column=i).fill = AMBER
        ws.row_dimensions[r].height = 60
        r += 1
    widths(ws, [30,28,20,15,14,16,10,70,42,17,50,32])

# ---------------------------------------------------------------- DQ tab
ws = wb.create_sheet('DQ candidates')
ws['A1'] = 'DQ-candidates — gated in Part A triage, with the quoted phrase that fired the gate'; ws['A1'].font = TITLE
ws['A2'] = ('These names go into the DQ log so they are never re-screened cold. The quoted phrase is verbatim from the Grata description. '
            'Note: the residential gate is literal per the prompt — a firm listing residential alongside commercial/industrial fires it, '
            'so read the quote before treating a residential DQ as final.')
ws['A2'].font = Font(italic=True, size=9); ws['A2'].alignment = WRAP
header(ws, 4, ['Bucket','Company','Domain','HQ','Grata revenue est.','Grata employee est.','Gate(s) fired with quoted phrase','Source','Grata profile URL'])
r = 5
dq_n = 0
for b in ORDER:
    for c in CAND[b]['candidates']:
        if c['priority'] != 'DQ-candidate': continue
        dq_n += 1
        row = [b, c['name'], c.get('domain') or '[unknown]', c.get('hq') or '[unknown]',
               money(c.get('rev')), num(c.get('emp')), c.get('note',''), c.get('source',''), c.get('url','')]
        for i,x in enumerate(row,1):
            cell = ws.cell(row=r, column=i, value=x); cell.alignment = WRAP; cell.fill = RED; cell.border = THIN
        ws.row_dimensions[r].height = 55
        r += 1
widths(ws, [8,30,28,20,15,14,95,40,32])

# ---------------------------------------------------------------- Public comps
ws = wb.create_sheet('Public comps')
ws['A1'] = 'Part B — Grata public comps per bucket'; ws['A1'].font = TITLE
ws['A2'] = ('SCALE-DISCOUNT CAVEAT: every multiple below comes from a listed company at EMCOR / Quanta / Comfort Systems scale. '
            'These multiples DO NOT transfer to a $2-10M EBITDA private target without a significant size discount. '
            'A lower-middle-market electrical services business does not trade at 25x EBITDA. Treat this tab as directional context '
            'for pricing instinct, not as a pricing model. Amber rows are listed PARENTS of small subsidiaries that Grata surfaced as '
            '"comps" — they are not trade comparables and should be excluded from any read.')
ws['A2'].alignment = WRAP; ws['A2'].fill = AMBER
ws.row_dimensions[2].height = 58

OFF = re.compile(r'goldman sachs|vistra|kkr|heico|acco|ares|ofs capital|amadeus|kingsway|ferguson|sgs |sgs$|littelfuse', re.I)
r = 4
for b in ORDER:
    ws.cell(row=r, column=1, value='%s — %s' % (b, NAMES[b])).font = TITLE
    r += 1
    seeds = COMPS[b]['seeds']
    allc = []
    for s in seeds:
        ws.cell(row=r, column=1, value='Seed: ' + s['seed']).font = BOLD
        r += 1
        if s.get('note'):
            c = ws.cell(row=r, column=1, value='Note: ' + s['note']); c.alignment = WRAP; c.fill = AMBER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 30
            r += 1
        header(ws, r, ['Comp company (Grata)','Ticker','Listed entity','EV/EBITDA (LTM)','EV/Revenue (LTM)','Flag'])
        ws.freeze_panes = None
        r += 1
        for cp in s['comps']:
            off = bool(OFF.search(cp.get('o') or '')) or bool(OFF.search(cp.get('t') or ''))
            row = [cp['n'], cp.get('t') or '[unknown]', cp.get('o') or '[unknown]',
                   round(cp['ebitda'],2) if cp.get('ebitda') is not None else '[unknown]',
                   round(cp['rev'],2) if cp.get('rev') is not None else '[unknown]',
                   'listed parent, not a trade comparable' if off else '']
            for i,x in enumerate(row,1):
                cell = ws.cell(row=r, column=i, value=x); cell.alignment = WRAP; cell.border = THIN
                if off: cell.fill = AMBER
            if cp.get('ebitda') is not None and not off:
                allc.append((cp['n'], cp['ebitda'], cp.get('rev')))
            r += 1
        gs = s.get('grata_stats') or {}
        if gs:
            txt = 'Grata aggregate for this seed: median EV/EBITDA %s · median EV/Rev %s · min EV/EBITDA %s · max EV/EBITDA %s' % (
                round(gs.get('ev_ebitda_median'),2) if gs.get('ev_ebitda_median') is not None else '[unknown]',
                round(gs.get('ev_rev_median'),2) if gs.get('ev_rev_median') is not None else '[unknown]',
                round(gs.get('ev_ebitda_min'),2) if gs.get('ev_ebitda_min') is not None else '[unknown]',
                round(gs.get('ev_ebitda_max'),2) if gs.get('ev_ebitda_max') is not None else '[unknown]')
            ws.cell(row=r, column=1, value=txt).font = Font(italic=True, size=9)
            r += 1
        r += 1
    # union median across seeds, trade-relevant comps only
    if allc:
        seen = {}
        for n,e,v_ in allc: seen[n] = (e, v_)
        eb = [e for e,_ in seen.values() if e is not None]
        rv = [v_ for _,v_ in seen.values() if v_ is not None]
        ws.cell(row=r, column=1, value='UNION across seeds (trade-relevant comps only, de-duplicated): %d comps · median EV/EBITDA %.2fx · median EV/Rev %.2fx'
                % (len(seen), statistics.median(eb), statistics.median(rv))).font = BOLD
        r += 2
widths(ws, [46,16,34,17,17,34])

# ---------------------------------------------------------------- Market cuts
ws = wb.create_sheet('Market cuts')
ws['A1'] = 'Part C — market cuts per bucket (counts only, page 1 total from each filtered search)'; ws['A1'].font = TITLE
ws['A2'] = ('Counts are the size of the Grata universe matching each bucket\'s primary keyword search plus the common filters, '
            'varying ONE dimension at a time. These are market-structure counts, not candidate lists — they include companies already in the master. '
            'EBITDA proxy: $2-10M EBITDA maps to the $5-50M revenue box per the system doc.')
ws['A2'].alignment = WRAP; ws['A2'].font = Font(italic=True, size=9)
ws['A3'] = 'Seller Intent: ' + CUTS['_meta']['seller_intent']
ws['A4'] = 'Hold period: ' + CUTS['_meta']['hold_period_signal']
ws['A5'] = 'Unknown revenue: ' + CUTS['_meta']['unknown_revenue_note']
for rr in (3,4,5):
    ws.cell(row=rr, column=1).fill = AMBER; ws.cell(row=rr, column=1).alignment = WRAP

r = 7
header(ws, r, ['Bucket','Bucket name','Total (all filters, no cut)',
               'Rev <$5M (nurture zone)','Rev $5-15M','Rev $15-30M','Rev $30-50M','Rev $50M+ (above box)',
               'IN BOX $5-50M','Unknown-revenue pool',
               'Northeast','Southeast','Midwest','Texas + South Central','West',
               'Bootstrapped','PE-backed (consolidator gauge)','Public / public subsidiary',
               'founder_age_signal','hold_period_signal','Seller Intent High'])
r += 1
for b in ORDER:
    d = CUTS[b]; rv = d['revenue']; rg = d['region']; ow = d['ownership']; ex = d['exit']
    row = [b, NAMES[b], d['total'], rv['<5M'], rv['5-15M'], rv['15-30M'], rv['30-50M'], rv['50M+'],
           d['in_box_5_50M'], d['unknown_revenue_pool'],
           rg['Northeast'], rg['Southeast'], rg['Midwest'], rg['Texas+South Central'], rg['West'],
           ow['bootstrapped'], ow['pe_backed'], ow['public+public_sub'],
           ex['founder_age_signal'],
           'unavailable' if not ex.get('hold_period_signal') else ex['hold_period_signal'],
           'unavailable']
    for i,x in enumerate(row,1):
        cell = ws.cell(row=r, column=i, value=x); cell.alignment = TOP; cell.border = THIN
        if i == 9: cell.font = BOLD
        if i in (20,21): cell.fill = AMBER
    r += 1
tots = ['TOTAL','']
for key in range(2, 21):
    pass
for i,fn in enumerate([lambda d: d['total'], lambda d: d['revenue']['<5M'], lambda d: d['revenue']['5-15M'],
                       lambda d: d['revenue']['15-30M'], lambda d: d['revenue']['30-50M'], lambda d: d['revenue']['50M+'],
                       lambda d: d['in_box_5_50M'], lambda d: d['unknown_revenue_pool'],
                       lambda d: d['region']['Northeast'], lambda d: d['region']['Southeast'],
                       lambda d: d['region']['Midwest'], lambda d: d['region']['Texas+South Central'],
                       lambda d: d['region']['West'], lambda d: d['ownership']['bootstrapped'],
                       lambda d: d['ownership']['pe_backed'], lambda d: d['ownership']['public+public_sub'],
                       lambda d: d['exit']['founder_age_signal']]):
    tots.append(sum(fn(CUTS[b]) for b in ORDER))
tots += ['unavailable','unavailable']
for i,x in enumerate(tots,1):
    c = ws.cell(row=r, column=i, value=x); c.font = BOLD
r += 2
ws.cell(row=r, column=1, value='Note: bucket universes overlap (a firm can satisfy more than one keyword profile), so the TOTAL row is a sum of overlapping sets, not a distinct company count.').font = Font(italic=True, size=9)
widths(ws, [8,30]+[16]*19)


# ---------------------------------------------------------------- Wave 2 tab
ws = wb.create_sheet('Wave 2 xref')
ws['A1'] = 'Wave 2 - Inven x Grata cross-reference: net-new companies in no existing list'; ws['A1'].font = TITLE
ws['A2'] = ('Broader, lower-precision pool than the wave-1 bucket tabs. Every row survived deterministic dedupe against all lists (wave 1 included) '
            'and an adversarial verification pass. Revenue/employee figures are single-platform ESTIMATES (Inven or Grata per the Source column) and '
            'disagree across platforms by up to 4.9x. Red = DQ-candidate (do not ingest into DQ log without the quoted-phrase check), amber = review/age/nurture flag.')
ws['A2'].alignment = WRAP; ws['A2'].font = Font(italic=True, size=9)
W2COLS = ['Bucket','Company','Domain','HQ','Revenue est.','Employee est.','Ownership','Year founded','Description (first 300 chars)','Source','Provisional priority','Gate / flag notes','Profile URL']
header(ws, 4, W2COLS)
r = 5
for b in ORDER:
    for c in sorted(W2['perBucket'][b]['candidates'], key=lambda x: (PRI_ORDER.get(x['priority'],9), -(x.get('rev') or 0))):
        row = [b, c['name'], c.get('domain') or '[unknown]', c.get('hq') or '[unknown]', money(c.get('rev')), num(c.get('emp')),
               c.get('own') or '[unknown]', num(c.get('yr')), (c.get('desc') or '')[:300], c.get('source',''), c['priority'], c.get('note','')[:500], c.get('url','')]
        for i,x in enumerate(row,1):
            cell = ws.cell(row=r, column=i, value=x); cell.alignment = WRAP; cell.border = THIN
        if c['priority'] == 'DQ-candidate':
            for i in range(1, len(W2COLS)+1): ws.cell(row=r, column=i).fill = RED
        elif 'REVIEW' in c.get('note','') or 'AGE FLAG' in c.get('note','') or c['priority'] == 'Nurture-candidate':
            for i in range(1, len(W2COLS)+1): ws.cell(row=r, column=i).fill = AMBER
        ws.row_dimensions[r].height = 52
        r += 1
widths(ws, [7,30,26,20,13,12,15,10,64,34,16,55,30])

# ---------------------------------------------------------------- Wave 3 tab
ws = wb.create_sheet('Wave 3 recovery')
ws['A1'] = 'Wave 3 - revenue-only recovery: in-box companies the employee proxy hid'; ws['A1'].font = TITLE
ws['A2'] = ('Same wave-1 primary keyword searches, revenue $5-50M, NO employee filter, deduped against every list and waves 1-2. '
            'Headcount shown for information only - it is not a criterion. Red = DQ-candidate, amber = review/age flag.')
ws['A2'].alignment = WRAP; ws['A2'].font = Font(italic=True, size=9)
header(ws, 4, W2COLS)
r = 5
for b in ORDER:
    for c in sorted(W3['perBucket'][b]['survivors'], key=lambda x: (PRI_ORDER.get(x['priority'],9), -(x.get('rev') or 0))):
        row = [b, c['name'], c.get('domain') or '[unknown]', c.get('hq') or '[unknown]', money(c.get('rev')), num(c.get('emp')),
               c.get('own') or '[unknown]', num(c.get('yr')), (c.get('desc') or '')[:300], c.get('source',''), c['priority'], c.get('note','')[:500], c.get('url','')]
        for i,x in enumerate(row,1):
            cell = ws.cell(row=r, column=i, value=x); cell.alignment = WRAP; cell.border = THIN
        if c['priority'] == 'DQ-candidate':
            for i in range(1, len(W2COLS)+1): ws.cell(row=r, column=i).fill = RED
        elif 'REVIEW' in c.get('note','') or 'AGE FLAG' in c.get('note',''):
            for i in range(1, len(W2COLS)+1): ws.cell(row=r, column=i).fill = AMBER
        ws.row_dimensions[r].height = 52
        r += 1
widths(ws, [7,30,26,20,13,12,15,10,64,34,16,55,30])

# ---------------------------------------------------------------- Conferences tab
CONF = json.load(open('scripts/xref/conferences_forward.json'))
ws = wb.create_sheet('Conferences')
ws['A1'] = 'Conference calendar - FORWARD ONLY: events starting %s to %s' % tuple(CONF['window']); ws['A1'].font = TITLE
ws['A2'] = ('Which upcoming events our targets attend. 2027 editions mostly have empty rosters today, so the prior-edition column carries the '
            'attendance evidence (e.g. 28 P2s attended EASA 2026 - book EASA 2027). Rosters are free to re-pull ~2-3 months before each event. '
            'NETA PowerTest, LPI, TSDOS and USMA 2027 editions are not yet in Grata\'s index (gap, not cancellation). '
            'Attendance is corroborating trade evidence, never a bucket assignment (sourcing rule 1).')
ws['A2'].alignment = WRAP; ws['A2'].font = Font(italic=True, size=9)
AUDIT = json.load(open('scripts/xref/conference_roster_audit.json'))
def _audit_for(name, grata_url):
    import re as _re
    # 1. match by event id embedded in the audit's future_event string
    m = _re.search(r'/([A-Z0-9]{8})\?', grata_url or '') or _re.search(r'lists/([A-Z0-9]{8})', grata_url or '')
    if m:
        for a in AUDIT['events']:
            if m.group(1) in a['future_event']: return a
    # 2. fuzzy: squash to alphanumerics and test containment on distinctive stems
    def sq(x): return _re.sub(r'[^a-z0-9]', '', x.lower())
    key = sq(_re.sub(r'\(.*?\)|CURRENT roster|- .*$', '', name))
    for a in AUDIT['events']:
        ak = sq(_re.sub(r'\(.*?\)|- CURRENT roster.*$|, [A-Za-z ]+\(', '', a['future_event']))
        if ak and key and (ak in key or key in ak or ak[:24] == key[:24]): return a
    # 3. token overlap on year + one distinctive word
    yr = _re.search(r'20\d\d', name)
    words = set(w for w in _re.findall(r'[a-z]{5,}', name.lower()) if w not in ('conference','convention','trade','annual','association','exposition'))
    for a in AUDIT['events']:
        if yr and yr.group(0) not in a['future_event']: continue
        aw = set(_re.findall(r'[a-z]{5,}', a['future_event'].lower()))
        if len(words & aw) >= 2: return a
    return None
header(ws, 4, ['Event','Dates','Location','VERDICT (roster audit)','% target-profile in prior roster','Already-known in room','Net-new discoveries','P2s attending now','Prior-edition evidence','Audit caveat','Event site','Grata attendee list'])
r = 5
for e in CONF['events']:
    when = '%s to %s' % (e.get('start_date') or '?', e.get('end_date') or '?')
    loc = ', '.join(x for x in [e.get('city'), e.get('state')] if x)
    p2names = '; '.join('%s (%s, %s)' % (a['name'], a['bucket'], a['wave']) for a in e['p2_attending'])
    pe = e.get('prior_edition_evidence')
    petxt = ('%s: %s P2, %s master' % (pe.get('edition'), pe.get('p2_count'), pe.get('master_count'))) if pe else '-'
    a = _audit_for(e['name'], e.get('grata_url'))
    comp = (a or {}).get('composition') or {}
    tp = comp.get('target_profile'); rs = (a or {}).get('roster_size')
    pct = ('%.0f%%' % (100.0*tp/rs)) if (tp is not None and rs) else '-'
    row = [e['name'], when, loc, (a or {}).get('verdict') or '-', pct,
           comp.get('already_known') if comp else '-', (a or {}).get('net_new_count') if a else '-',
           len(e['p2_attending']), petxt, ('; '.join((a or {}).get('caveats') or [])[:200]) if a else '',
           e.get('source_link') or '', e.get('grata_url') or '']
    for i2,x in enumerate(row,1):
        cell = ws.cell(row=r, column=i2, value=x); cell.alignment = WRAP; cell.border = THIN
    v = (a or {}).get('verdict') or ''
    if v == 'STRONG GO':
        for i2 in range(1, 13): ws.cell(row=r, column=i2).fill = AMBER
    elif v.startswith('SKIP'):
        for i2 in range(1, 13): ws.cell(row=r, column=i2).fill = RED
    ws.row_dimensions[r].height = 42
    r += 1
r += 1
ws.cell(row=r, column=1, value=('ROSTER AUDIT: prior-edition rosters were pulled in full (free) and every company classified. %d net-new target-profile companies were '
    'discovered from rosters alone (largest hauls: EASA 54, NECA family ~45, NRECA/Expo ~25) - names with rationale in scripts/xref/conference_roster_audit.json; '
    'they need description-driven triage before entering any list (rule 1) and are NOT deduped against the Dealbuff list (rule 7). '
    'NECA rosters are exhibitor halls, so their low target-percent understates owner density on the floor. '
    'Grata now hard-caps list reads at 500 rows (large-company-first), so mega-show tails are systematically unseen.') % AUDIT['net_new_total']).font = Font(size=9)
ws.cell(row=r, column=1).alignment = WRAP; ws.cell(row=r, column=1).fill = AMBER
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12); ws.row_dimensions[r].height = 60
r += 2
ws.cell(row=r, column=1, value='Discarded off-window or off-thesis events (%d - reasons in scripts/xref/conferences_forward.json): %s' % (
    len(CONF['discarded_events']), '; '.join(x['name'] for x in CONF['discarded_events'][:14]))).font = Font(italic=True, size=9)
ws.cell(row=r, column=1).alignment = WRAP
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12); ws.row_dimensions[r].height = 36
widths(ws, [38,19,16,14,12,11,11,10,24,46,24,24])

# ---------------------------------------------------------------- M&A intel tab
MA = json.load(open('scripts/xref/ma_intel.json'))
ws = wb.create_sheet('M&A intel')
ws['A1'] = 'M&A intelligence - transactions, live deals, bankers, buyers (Grata, as of %s)' % MA['as_of']; ws['A1'].font = TITLE
r = 3
def block(title):
    global r
    ws.cell(row=r, column=1, value=title).font = TITLE; r += 1
def note(txt, fill=None, h=40):
    global r
    c = ws.cell(row=r, column=1, value=txt); c.alignment = WRAP; c.font = Font(size=9)
    if fill: c.fill = fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); ws.row_dimensions[r].height = h
    r += 1

block('1. CLOSED TRANSACTIONS')
note('Universe: %d US electrical-contractor deals in 5 years; %d US industrial-repair deals in 3 years. %s'
     % (MA['transactions']['universe']['electrical_contractors_5yr_US'], MA['transactions']['universe']['industrial_repair_3yr_US'], MA['transactions']['multiples_note']), AMBER, 55)
header(ws, r, ['Target','Buyer','Date','EV','EV/EBITDA','EV/Rev','Note','']); r += 1
for t in MA['transactions']['disclosed_multiples']:
    row = [t['target'], t['buyer'], t['date'], '$%.1fM' % (t['ev']/1e6), t.get('ev_ebitda') or '', t.get('ev_rev') or '', t.get('note',''), '']
    for i,x in enumerate(row,1): ws.cell(row=r, column=i, value=x).alignment = WRAP
    r += 1
r += 1
ws.cell(row=r, column=1, value='Notable recent deals (consolidator activity):').font = BOLD; r += 1
for t in MA['transactions']['notable_recent']:
    note('%s <- %s (%s, %s): %s' % (t['target'], t['buyer'], t['date'], t['type'], t['note']), None, 26)
r += 1
block('2. LIVE DEALS - businesses on the market now')
header(ws, r, ['Mandate','Revenue','EBITDA','Ask / multiple','Advisor','Verified','Read','']); r += 1
for m in MA['live_deals']:
    ask = ('$%.1fM' % (m['ask']/1e6) if m.get('ask') else '') + ((' / %.1fx' % m['ask_multiple']) if m.get('ask_multiple') else '')
    row = [m['name'], '$%.1fM' % (m['rev']/1e6), '$%.2fM' % (m['ebitda']/1e6), ask or '[unknown]', m.get('advisor',''), 'yes' if m.get('verified') else '', m['read'], '']
    for i,x in enumerate(row,1):
        cell = ws.cell(row=r, column=i, value=x); cell.alignment = WRAP
    if 'IN BOX' in m['read']:
        for i in range(1,9): ws.cell(row=r, column=i).fill = AMBER
    ws.row_dimensions[r].height = 40
    r += 1
note(MA['ask_multiple_read'], AMBER, 30)
r += 1
block('3. BANKERS - who sells these businesses')
header(ws, r, ['Firm','Domain / evidence','Why relevant','','','','','']); r += 1
for b in MA['bankers']['sector_specialists']:
    for i,x in enumerate([b['name'], b['domain'] + ' (' + b['hq'] + ')', b['why'],'','','','',''],1): ws.cell(row=r, column=i, value=x).alignment = WRAP
    ws.row_dimensions[r].height = 28; r += 1
for b in MA['bankers']['deal_evidenced']:
    for i,x in enumerate([b['name'], 'deal-evidenced', b['evidence'],'','','','',''],1): ws.cell(row=r, column=i, value=x).alignment = WRAP
    ws.row_dimensions[r].height = 24; r += 1
r += 1
block('4. BUYERS - the consolidator map')
note(MA['buyers']['read'], AMBER, 45)
ws.cell(row=r, column=1, value='PE sponsors with LMM electrical mandates (%d total; top by matching portfolio cos):' % MA['buyers']['pe_universe_count']).font = BOLD; r += 1
for b in MA['buyers']['pe_sponsors_lmm']:
    for i,x in enumerate([b['name'], 'portfolio matches: %d' % b['matching_portfolio'], b.get('note',''),'','','','',''],1): ws.cell(row=r, column=i, value=x).alignment = WRAP
    ws.row_dimensions[r].height = 22; r += 1
ws.cell(row=r, column=1, value='Strategic acquirers (%d in universe; most active):' % MA['buyers']['strategic_universe_count']).font = BOLD; r += 1
for b in MA['buyers']['strategics']:
    for i,x in enumerate([b['name'], 'portfolio matches: %d' % b['matching_portfolio'], b.get('note',''),'','','','',''],1): ws.cell(row=r, column=i, value=x).alignment = WRAP
    ws.row_dimensions[r].height = 22; r += 1
widths(ws, [34,30,60,14,11,10,70,4])

wb.save(OUT)
print('wrote', OUT)
print('DQ rows:', dq_n, '| new candidates:', tot['n'])

# ---------------------------------------------------------------- Contacts
# Owner-level contact enrichment (Inven get_company_contacts, Grata fallback).
# Client needs three fields only: owner email, mobile phone, postal address.
import os
_CPATH = 'scripts/contacts/owner_contacts_wave123.json'
_MPATH = 'scripts/contacts/mainline_map.json'
if os.path.exists(_CPATH):
    CON = json.load(open(_CPATH))
    ADDR = json.load(open(_MPATH))['companies'] if os.path.exists(_MPATH) else {}

    def _best_email(c):
        # verified professional beats verified personal beats anything else
        es = c.get('emails') or []
        rank = lambda e: (not e.get('is_verified'), e.get('type') != 'professional')
        es = sorted(es, key=rank)
        return es[0]['email'] if es else None

    def _phones_by_type(c, t):
        return [p['number'] for p in (c.get('phones') or []) if p.get('type') == t]

    rows = []
    for dom, v in CON['companies'].items():
        cs = v.get('contacts') or []
        if not cs:
            rows.append({'dom': dom, 'v': v, 'c': None, 'mob': [], 'dd': [], 'off': [], 'em': None})
            continue
        for c in cs:
            rows.append({'dom': dom, 'v': v, 'c': c,
                         'mob': _phones_by_type(c, 'mobile'),
                         'dd':  _phones_by_type(c, 'direct dial'),
                         'off': _phones_by_type(c, 'office'),
                         'em':  _best_email(c)})
    # most actionable first: has mobile, then direct dial, then email only
    rows.sort(key=lambda r: (not r['mob'], not r['dd'], not r['em'],
                             str(r['v'].get('priority') or 'P9'), r['dom']))

    ws = wb.create_sheet('Contacts')
    ws.cell(row=1, column=1, value='OWNER CONTACTS - enrichment output').font = TITLE
    n_mob = sum(1 for r in rows if r['mob'])
    ws.cell(row=2, column=1, value=(
        '%d companies attempted, %d returned an owner-level contact, %d carry a MOBILE number. '
        'Inven contact credits spent: %d. Phone types are as the provider labelled them - "mobile" is the '
        'cell, "direct dial" is a personal desk line, "office" is the company mainline and is NOT a cell. '
        'Misses cost nothing: a title-filter miss means the provider held only non-owner staff.'
        % (CON['attempted'], CON['with_contacts'], n_mob, CON['credits_used'])
    )).alignment = WRAP
    ws.row_dimensions[2].height = 42
    r = 4
    header(ws, r, ['Company', 'Domain', 'Owner / contact', 'Title', 'MOBILE', 'Direct dial',
                   'Office (mainline - not a cell)', 'Email', 'Postal address', 'Bucket',
                   'Campaign', 'Priority', 'Status / note'])
    r += 1
    for x in rows:
        v, c = x['v'], x['c']
        addr = (ADDR.get(x['dom']) or {}).get('street_address') or '[unknown]'
        vals = [v.get('name') or '[unknown]', x['dom'],
                (c or {}).get('name') or '[unknown]', (c or {}).get('title') or '[unknown]',
                ', '.join(x['mob']) or '[unknown]', ', '.join(x['dd']) or '[unknown]',
                ', '.join(x['off']) or '', x['em'] or '[unknown]', addr,
                v.get('bucket') or '', v.get('campaign') or '', v.get('priority') or '',
                v.get('error') or ('REPLIED - active conversation' if v.get('replied') else 'contact found')]
        for i, val in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=val).alignment = WRAP
        if x['mob']:
            ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='E2EFDA')  # green = actionable cell
        elif x['off'] and not x['dd']:
            for i in range(1, 14): ws.cell(row=r, column=i).fill = AMBER  # mainline only - needs cell pass
        ws.row_dimensions[r].height = 26
        r += 1
    widths(ws, [30, 26, 20, 30, 18, 18, 20, 32, 44, 8, 22, 8, 34])
    wb.save(OUT)
    print('Contacts tab:', len(rows), 'rows |', n_mob, 'with mobile')
