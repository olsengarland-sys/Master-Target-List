"""Export the 627 unvalidated expansion companies for the client to screen.

These reached the Grata/Inven expansion but never went through the campaign
validation pass, so they carry no bucket verdict and no size check. Full
descriptions come from the JSON stores, not the workbook: the workbook holds a
300-char truncation, and truncation is what caused 47 unresolved verdicts in
the last validation round.
"""
import json, re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

U = '/root/.claude/uploads/cdc2f6cb-7804-5be0-bfc8-f894363d3735/'
M = U + 'cee59e13-Hunter_Power_Master_Target_List_20260825_v3.xlsx'
N = U + 'd2024f6e-Hunter_Power_New_Targets_for_Campaigns_20260901.xlsx'
G = U + '48c5eb96-Hunter_Power_Grata_Expansion_20260831.xlsx'
R = '/home/user/Master-Target-List/'

def dom(v):
    d = str(v or '').lower()
    d = re.sub(r'^https?://', '', d); d = re.sub(r'^www\.', '', d)
    return d.split('/')[0] or None

def grab(path, sheets, hr, key):
    out = set(); wb = load_workbook(path, read_only=True, data_only=True)
    for s in sheets:
        if s not in wb.sheetnames: continue
        ws = wb[s]; hdr = [c.value for c in next(ws.iter_rows(min_row=hr, max_row=hr))]
        idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
        if key not in idx: continue
        for r in ws.iter_rows(min_row=hr + 1, values_only=True):
            if any(r) and r[idx[key]]: out.add(dom(r[idx[key]]))
    wb.close(); return out

CAMP = ['NEW additions (138)', 'Commercial & Industrial Ctrs', 'Motor & Generator Companies',
        'Utility Line & Substation', 'Mission critical power', 'Traffic Signals & Lighting',
        'Testing & Commissioning', 'Electrical Apparatus Servicing', 'EV Charging',
        'Lightning Protection', 'Electrical Equipment Mfg', 'Power Systems Engineering']
master = grab(M, ['All 1922 master', 'NEW - Add to batch 1 or 2 (11)', 'NEW - Add to batch 2 or 3 (51)'], 2, 'Website')
camp = grab(N, CAMP, 2, 'Domain'); campdq = grab(N, ['Remove - DQ'], 2, 'Domain')
grata = grab(G, ['T1','T2','T3','T4','T5','T6','S7','S8','S9','S10','Wave 2 xref','Wave 3 recovery'], 4, 'Domain')
gratadq = grab(G, ['DQ candidates'], 4, 'Domain')
target = grata - camp - master - campdq - gratadq

# full records (untruncated descriptions) out of the JSON stores
rec = {}
def soak(store, wave):
    for b, blk in store.items():
        cands = blk.get('candidates') if isinstance(blk, dict) else None
        for c in (cands or []):
            d = dom(c.get('domain'))
            if d and d not in rec:
                rec[d] = dict(c, bucket=b, wave=wave)
soak(json.load(open(R + 'scripts/candidates.json')), 'Wave 1 (bucket search)')
soak(json.load(open(R + 'scripts/xref/wave2.json'))['perBucket'], 'Wave 2 (Inven x Grata)')
soak(json.load(open(R + 'scripts/xref/wave3_recovery.json'))['perBucket'], 'Wave 3 (revenue-only)')

C = json.load(open(R + 'scripts/contacts/owner_contacts_wave123.json'))['companies']
A = json.load(open(R + 'scripts/contacts/addresses.json'))['companies']
QC = json.load(open(R + 'scripts/contacts/contact_qc_flags.json'))
FLAG = ({x['domain'] for x in QC['verify_before_outreach']} | {x['domain'] for x in QC['us_only_violations']}
        | {x['domain'] for x in QC['bad_domains']})
OWNER = re.compile(r'owner|president|chief executive|\bceo\b|founder|principal|proprietor|'
                   r'managing (director|partner|member)|general manager|\bpartner\b', re.I)

# DQ suggestions must be CONTEXT-AWARE. A naive keyword sweep flags "commercial,
# industrial, and residential" as residential, and a utility distribution
# contractor as underground civil - the exact failure the validation was meant
# to fix. So a gate only fires when the pattern describes the SUBJECT of the
# description AND no core electrical-services signal is present.
IN_THESIS = re.compile(
    r'electrical contract|electrical service|electrical construct|switchgear|circuit breaker|'
    r'transformer|motor (repair|rewind|shop)|rewind|substation|utility distribution|'
    r'transmission (and|&) distribution|power line|line contract|electrical testing|'
    r'acceptance testing|commissioning|relay|arc.flash|generator (service|maintenance)|'
    r'ups (service|maintenance)|critical power|traffic signal|lightning protection|'
    r'electrical maintenance|electrical installation|electrical infrastructure', re.I)

# each: (pattern, label, whether an in-thesis signal cancels it)
GATES = [
 (re.compile(r'cathodic protection|corrosion control|corrosion engineer', re.I),
  'cathodic/corrosion - validated DQ pattern', False),
 (re.compile(r'^\s*\S[^.]{0,120}?\b(solar|photovoltaic)\b[^.]{0,80}(systems|installer|installation|developer|epc)', re.I),
  'solar may be the primary business', True),
 (re.compile(r'\bresidential\b', re.I), 'residential', True),
 (re.compile(r'\b(is|as)\s+(a|an)\s+[^.]{0,60}(distributor|wholesaler|supply house|reseller|'
             r"manufacturer's (sales )?representative)", re.I),
  'distributor / rep firm, not a services business', False),
 (re.compile(r'safety (products|training) (company|provider|firm)|provides .{0,40}safety training', re.I),
  'safety products/training', True),
 (re.compile(r'\b(is|provides)\b[^.]{0,60}\b(telecommunications?|fiber optic|fibre|cell tower)\b', re.I),
  'telecom / fibre', True),
 (re.compile(r'\b(is|provides)\b[^.]{0,60}\b(directional drilling|underground civil|excavation|trenching)\b', re.I),
  'underground civil', True),
 (re.compile(r'sign (fabricat|manufactur)|highway striping|guardrail', re.I), 'signage / highway', True),
 (re.compile(r'\bstaffing\b|recruit(ing|ment)|placement agency', re.I), 'staffing', True),
 (re.compile(r'\b(is|develops|provides)\b[^.]{0,40}\b(software|saas platform)\b', re.I), 'software', True),
]

def truncated(txt):
    """Some stored descriptions are cut mid-sentence; a gate must not judge on those."""
    if not txt:
        return True
    t = txt.strip()
    return not t.endswith(('.', '!', '?')) or len(t) < 90

def suggest(txt, name):
    """Return DQ labels only where the pattern plausibly describes this company."""
    if truncated(txt):
        return 'DESCRIPTION INCOMPLETE - decide from the company website'
    s = f'{name or ""}. {txt or ""}'
    thesis = bool(IN_THESIS.search(s))
    hits = []
    for rx, lab, cancelable in GATES:
        if not rx.search(s):
            continue
        if cancelable and thesis:
            # residential is the special case: only DQ when it is the ONLY market named
            if lab == 'residential' and not re.search(r'commercial|industrial|institutional|utility|municipal', s, re.I):
                hits.append('residential only - no commercial/industrial named')
            continue
        hits.append(lab)
    return '; '.join(hits)

rows = []
for d in sorted(target):
    r = rec.get(d, {})
    v = C.get(d); c = {}
    if v and v.get('contacts'):
        cs = v['contacts']; best = next((x for x in cs if OWNER.search(x.get('title') or '')), cs[0])
        ph = lambda t: ', '.join(p['number'] for p in (best.get('phones') or []) if p.get('type') == t)
        es = sorted((best.get('emails') or []), key=lambda e: (not e.get('is_verified'), e.get('type') != 'professional'))
        c = {'n': best.get('name'), 't': best.get('title'), 'm': ph('mobile'),
             'e': es[0]['email'] if es else None}
    a = A.get(d) or {}
    desc = r.get('desc')
    rows.append({
        'company': r.get('name'), 'domain': d, 'hq': r.get('hq'),
        'bucket': r.get('bucket'), 'wave': r.get('wave'),
        'revenue': r.get('rev'), 'employees': r.get('emp'), 'founded': r.get('yr'),
        'ownership': r.get('own'), 'priority': r.get('priority'),
        'description': desc, 'source': r.get('source'),
        'suggest': suggest(desc, r.get('name')),
        'contact': c.get('n'), 'title': c.get('t'), 'email': c.get('e'), 'mobile': c.get('m'),
        'address': a.get('street_address') if a.get('mailable') else None,
        'qc': 'YES' if d in FLAG else '',
    })
# most decidable first: has a description, then has contact data
rows.sort(key=lambda x: (not x['description'], not x['mobile'], str(x['company'] or '')))

HDR = PatternFill('solid', fgColor='1F3864'); HF = Font(color='FFFFFF', bold=True, size=10)
AMBER = PatternFill('solid', fgColor='FFF2CC'); RED = PatternFill('solid', fgColor='FCE4E4')
GREEN = PatternFill('solid', fgColor='E2EFDA'); WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook(); ws = wb.active; ws.title = 'READ ME'
ws.cell(row=1, column=1, value='627 unvalidated companies - for screening').font = Font(bold=True, size=14)
txt = [
 '', 'These reached the Grata/Inven expansion, are absent from Master Target List v3, and were NOT',
 'DQd by anyone - but they never went through the campaign validation pass either. So they carry',
 'no bucket verdict and no size check.', '',
 'Where they sit: the expansion produced 1,755 candidates. 729 reached a campaign sheet, 315 were',
 'DQd by the campaign validation and 84 in the expansion itself. These 627 are the untouched remainder.', '',
 'EXPECT A HIGH REJECT RATE. Validation threw out 315 of the 1,026 it examined - about 31%. If these',
 'behave the same way, roughly 200 of the 627 are DQ material.', '',
 'HOW TO WORK IT', '',
 'Three empty columns at the right are for you: VERDICT, CORRECT BUCKET/CAMPAIGN, REASON.',
 'Sorted so the decidable ones come first - rows with a description, then rows with contact data.', '',
 'The "Suggested DQ" column pre-flags rows matching patterns the last validation actually fired on',
 '(cathodic protection, solar, residential, distribution, telecom, underground civil, signage,',
 'staffing, software, safety training). It is a first read, not a decision - nothing was removed.', '',
 'CLASSIFY FROM THE DESCRIPTION, NOT THE SEARCH THAT FOUND IT. The Source column shows which keyword',
 'surfaced each company; filing by that produced 0-7% validity in the specialist buckets last time.',
 'A company found by a lightning-protection search that reads as a commercial contractor is T5.', '',
 'Descriptions here are FULL, not the 300-character truncation in the expansion workbook - the',
 'truncation is what left 47 verdicts unresolved last round.', '',
 'Contact data is joined on where held, so a company you keep is ready to work immediately.',
]
r2 = 3
for t in txt:
    ws.cell(row=r2, column=1, value=t); r2 += 1
ws.column_dimensions['A'].width = 118

ws = wb.create_sheet('TO VALIDATE (627)')
cols = [('Company',30),('Domain',26),('HQ',22),('Bucket (provisional)',14),('Found in',22),
        ('Revenue est.',14),('Employees',11),('Founded',9),('Ownership',15),('Provisional priority',16),
        ('FULL description',80),('Source (search that found it)',34),('Suggested DQ - check',30),
        ('Owner / contact',18),('Title',24),('Email',28),('MOBILE',18),('Postal address',36),('QC flag',9),
        ('VERDICT (keep / DQ / nurture)',26),('CORRECT BUCKET / CAMPAIGN',26),('REASON',40)]
for i,(h,w) in enumerate(cols,1):
    c = ws.cell(row=1, column=i, value=h); c.fill = HDR; c.font = HF; c.alignment = WRAP
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'C2'
keys = ['company','domain','hq','bucket','wave','revenue','employees','founded','ownership','priority',
        'description','source','suggest','contact','title','email','mobile','address','qc']
for n,row in enumerate(rows,2):
    for i,k in enumerate(keys,1):
        ws.cell(row=n, column=i, value=row.get(k)).alignment = WRAP
    for i in (20,21,22):
        ws.cell(row=n, column=i).fill = GREEN     # your columns to fill
    if row['suggest']:
        ws.cell(row=n, column=13).fill = AMBER
    if row['qc']:
        ws.cell(row=n, column=19).fill = RED
    ws.row_dimensions[n].height = 30

OUT = '/home/user/Master-Target-List/Hunter_Power_627_To_Validate_20260902.xlsx'
wb.save(OUT)
print('rows:', len(rows),
      '| with full description:', sum(1 for r in rows if r['description']),
      '| suggested DQ:', sum(1 for r in rows if r['suggest']),
      '| with mobile:', sum(1 for r in rows if r['mobile']),
      '| with email:', sum(1 for r in rows if r['email']))
print('wrote', OUT)
