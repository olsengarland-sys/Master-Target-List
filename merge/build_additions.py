"""Build ONE sheet of companies to add to Master Target List v3.

Two groups:
  A. NET-NEW - companies in the three later files whose domain is absent from
     master v3 entirely.
  B. NURTURE - master v3's own Nurture rows (already in the master, but parked
     below the EBITDA floor; the client wants them back in the target list).

DQ rows in the later files are excluded: they were removed deliberately and
must not be re-added.
"""
import re
from openpyxl import load_workbook

U = '/root/.claude/uploads/cdc2f6cb-7804-5be0-bfc8-f894363d3735/'
MASTER = U + 'cee59e13-Hunter_Power_Master_Target_List_20260825_v3.xlsx'
NEWT   = U + 'd2024f6e-Hunter_Power_New_Targets_for_Campaigns_20260901.xlsx'
VALID  = U + '6500def4-Hunter_Power_Validated_Campaigns_20260901.xlsx'
GRATA  = U + '48c5eb96-Hunter_Power_Grata_Expansion_20260831.xlsx'

def norm(d):
    """Normalize a website/domain to a bare lowercase host."""
    if not d:
        return None
    d = str(d).strip().lower()
    d = re.sub(r'^https?://', '', d)
    d = re.sub(r'^www\.', '', d)
    d = d.split('/')[0].split('?')[0].strip()
    return d or None

def rows(path, sheet, header_row):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr = [c.value for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
    out = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(r):
            continue
        out.append((r, idx))
    wb.close()
    return out

# ---------- master v3: the baseline ----------
known = set()
for r, idx in rows(MASTER, 'All 1922 master', 2):
    d = norm(r[idx['Website']])
    if d:
        known.add(d)
print('master v3 domains:', len(known))

nurture = []
for r, idx in rows(MASTER, 'Nurture (below floor)', 2):
    nurture.append({h: r[i] for h, i in idx.items()})
print('master v3 Nurture rows:', len(nurture))

# ---------- the three later files ----------
dq = set()
for path, sheet, hr in [(NEWT, 'Remove - DQ', 2), (VALID, 'Remove - DQ', 2)]:
    for r, idx in rows(path, sheet, hr):
        key = 'Domain' if 'Domain' in idx else None
        if key:
            d = norm(r[idx[key]])
            if d:
                dq.add(d)
for r, idx in rows(GRATA, 'DQ candidates', 4):
    for k in ('Domain', 'domain'):
        if k in idx:
            d = norm(r[idx[k]])
            if d:
                dq.add(d)
            break
print('DQ domains to exclude:', len(dq))

CAMPAIGN_SHEETS = ['NEW additions (138)', 'Commercial & Industrial Ctrs', 'Motor & Generator Companies',
                   'Utility Line & Substation', 'Mission critical power', 'Traffic Signals & Lighting',
                   'Testing & Commissioning', 'Electrical Apparatus Servicing', 'EV Charging',
                   'Lightning Protection', 'Electrical Equipment Mfg', 'Power Systems Engineering']
GRATA_SHEETS = ['T1','T2','T3','T4','T5','T6','S7','S8','S9','S10','Wave 2 xref','Wave 3 recovery']

cand = {}
def add(d, rec):
    if not d or d in known or d in dq:
        return
    if d in cand:
        cand[d]['sources'].add(rec['source'])
        for k, v in rec.items():
            if k != 'source' and v and not cand[d].get(k):
                cand[d][k] = v
    else:
        rec['sources'] = {rec.pop('source')}
        cand[d] = rec

for path, label in [(NEWT, 'New Targets for Campaigns (1 Sep)'), (VALID, 'Validated Campaigns (1 Sep)')]:
    wb = load_workbook(path, read_only=True); names = wb.sheetnames; wb.close()
    for sh in CAMPAIGN_SHEETS:
        if sh not in names:
            continue
        for r, idx in rows(path, sh, 2):
            g = lambda k: r[idx[k]] if k in idx else None
            add(norm(g('Domain')), {
                'source': label, 'company': g('Company'), 'hq': g('HQ'),
                'bucket': g('Bucket'), 'campaign': g('Correct campaign') or sh,
                'priority': g('Priority'), 'employees': g('Employee est.'),
                'revenue': g('Revenue est.'), 'founded': g('Year founded'),
                'ownership': g('Ownership'), 'description': g('reason') or g('Description'),
                'contact': g('Contact'), 'contact_title': g('Contact title'),
                'contact_email': g('Contact email'), 'size_read': g('Size read'),
            })

for sh in GRATA_SHEETS:
    for r, idx in rows(GRATA, sh, 4):
        g = lambda k: r[idx[k]] if k in idx else None
        add(norm(g('Domain')), {
            'source': 'Grata Expansion (31 Aug)', 'company': g('Company'),
            'hq': g('HQ (city, state)') or g('HQ'), 'bucket': g('Bucket') or sh,
            'campaign': None, 'priority': g('Provisional priority'),
            'employees': g('Grata employee est.') or g('Employee est.'),
            'revenue': g('Grata revenue est.') or g('Revenue est.'),
            'founded': g('Year founded'), 'ownership': g('Ownership'),
            'description': g('Grata description (first 300 chars)') or g('Description (first 300 chars)'),
            'contact': None, 'contact_title': None, 'contact_email': None, 'size_read': None,
        })

print('NET-NEW companies (not in master v3, not DQ):', len(cand))
import json, collections
print('by source:', collections.Counter(
    ' + '.join(sorted(v['sources'])) for v in cand.values()).most_common())
json.dump({'net_new': {k: {kk: (str(vv) if not isinstance(vv, (str, int, float, type(None))) else vv)
                           for kk, vv in v.items() if kk != 'sources'} | {'sources': sorted(v['sources'])}
                       for k, v in cand.items()},
           'nurture': [{k: (str(v) if not isinstance(v, (str, int, float, type(None))) else v)
                        for k, v in n.items()} for n in nurture]},
          open('merged.json', 'w'), indent=1, default=str)
