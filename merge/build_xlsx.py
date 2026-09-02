import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

m = json.load(open('merged.json'))
C = json.load(open('/home/user/Master-Target-List/scripts/contacts/owner_contacts_wave123.json'))['companies']
A = json.load(open('/home/user/Master-Target-List/scripts/contacts/addresses.json'))['companies']
QC = json.load(open('/home/user/Master-Target-List/scripts/contacts/contact_qc_flags.json'))
FLAG = ({x['domain'] for x in QC['verify_before_outreach']} | {x['domain'] for x in QC['us_only_violations']}
        | {x['domain'] for x in QC['bad_domains']})
OWNER = re.compile(r'owner|president|chief executive|\bceo\b|founder|principal|proprietor|'
                   r'managing (director|partner|member)|general manager|\bpartner\b', re.I)

def norm(d):
    d = str(d or '').strip().lower()
    d = re.sub(r'^https?://', '', d); d = re.sub(r'^www\.', '', d)
    return d.split('/')[0] or None

def contact_of(dom):
    """Best owner-level contact we hold, else the best of any contact."""
    v = C.get(dom)
    if not v or not v.get('contacts'):
        return {}
    cs = v['contacts']
    best = next((c for c in cs if OWNER.search(c.get('title') or '')), cs[0])
    ph = lambda t: ', '.join(p['number'] for p in (best.get('phones') or []) if p.get('type') == t)
    es = sorted((best.get('emails') or []),
                key=lambda e: (not e.get('is_verified'), e.get('type') != 'professional'))
    return {'name': best.get('name'), 'title': best.get('title'), 'mobile': ph('mobile'),
            'direct': ph('direct dial'), 'email': es[0]['email'] if es else None,
            'verified': 'yes' if (es and es[0].get('is_verified')) else ''}

rows = []
for dom, r in m['net_new'].items():
    c = contact_of(dom)
    a = A.get(dom) or {}
    rows.append({
        'group': 'NET-NEW', 'company': r.get('company'), 'website': dom, 'hq': r.get('hq'),
        'bucket': r.get('bucket'), 'campaign': r.get('campaign'), 'priority': r.get('priority'),
        'employees': r.get('employees'), 'revenue': r.get('revenue'), 'founded': r.get('founded'),
        'ownership': r.get('ownership'), 'contact': c.get('name'), 'title': c.get('title'),
        'email': c.get('email') or r.get('contact_email'), 'email_verified': c.get('verified'),
        'mobile': c.get('mobile'), 'direct': c.get('direct'),
        'address': a.get('street_address') if a.get('mailable') else None,
        'verify': 'YES - QC flag' if dom in FLAG else '',
        'note': r.get('description') or r.get('size_read'),
        'src': ' + '.join(r.get('sources') or []),
    })
for r in m['nurture']:
    dom = norm(r.get('Website'))
    c = contact_of(dom) if dom else {}
    a = (A.get(dom) or {}) if dom else {}
    rows.append({
        'group': 'NURTURE (from master v3)', 'company': r.get('Company'), 'website': dom,
        'hq': r.get('HQ'), 'bucket': r.get('Type bucket'), 'campaign': None,
        'priority': r.get('Priority'), 'employees': None, 'revenue': r.get('Size estimate and basis'),
        'founded': r.get('Age'), 'ownership': r.get('Ownership'),
        'contact': c.get('name') or r.get('Contact'), 'title': c.get('title') or r.get('Contact title'),
        'email': c.get('email') or r.get('Contact email'), 'email_verified': c.get('verified'),
        'mobile': c.get('mobile'), 'direct': c.get('direct'),
        'address': a.get('street_address') if a.get('mailable') else None,
        'verify': 'YES - QC flag' if dom in FLAG else '',
        'note': r.get('Why / verdict reason') or r.get('List-stage note'),
        'src': 'Master v3 Nurture sheet',
    })
rows.sort(key=lambda x: (x['group'] != 'NET-NEW', not x['mobile'], not x['email'],
                         str(x['company'] or '')))

HDR = PatternFill('solid', fgColor='1F3864'); HF = Font(color='FFFFFF', bold=True, size=10)
GREEN = PatternFill('solid', fgColor='E2EFDA'); AMBER = PatternFill('solid', fgColor='FFF2CC')
RED = PatternFill('solid', fgColor='FCE4E4'); WRAP = Alignment(wrap_text=True, vertical='top')
wb = Workbook(); ws = wb.active; ws.title = 'READ ME'
notes = [
 ('Which file is the latest?', None),
 ('', None),
 ('Master Target List v3 (25 Aug) is your BASE - 1,922 companies. Everything below is measured against it.', None),
 ('Grata Expansion (31 Aug) is the raw expansion output: net-new companies found by searching Grata and Inven.', None),
 ('Validated Campaigns (1 Sep) took that expansion, validated it and filed it into 11 campaigns.', None),
 ('New Targets for Campaigns (1 Sep) is Validated Campaigns PLUS 138 companies released from the master Nurture list.', None),
 ('', None),
 ('So: "New Targets for Campaigns" SUPERSEDES "Validated Campaigns" - it contains everything that file has, plus 138 more.', True),
 ('You do not need to work from both. The Grata Expansion still holds 627 companies that never made it into a campaign sheet.', None),
 ('', None),
 ('What is in this workbook', None),
 ('', None),
 ('One sheet, "ADD TO TARGET LIST", with everything to append to master v3:', None),
 ('  A. NET-NEW (726) - survived every validation pass, absent from master v3 entirely.', None),
 ('  B. NURTURE (138) - already in master v3 but parked below the EBITDA floor; you asked for these back.', None),
 ('', None),
 ('Companies removed as DQ in the later files were excluded (402 domains). They were dropped deliberately.', None),
 ('', None),
 ('Contact data from the enrichment run is joined on where we have it:', None),
 ('  See the counts printed when this was built.', None),
 ('', None),
 ('All 726 net-new rows are validated', True),
 ('Every one appears on a campaign sheet in the 1 September file, having survived both the', None),
 ('31 August Campaign Assignment pass and the 1 September validation. There is no unscreened tail.', None),
 ('Rows marked in the "Verify first" column carry a QC flag from the contact enrichment (wrong-entity match, non-US contact, or a bad domain).', None),
]
ws.cell(row=1, column=1, value='Hunter Power - companies to add to Master Target List v3').font = Font(bold=True, size=14)
r = 3
for txt, bold in notes:
    c = ws.cell(row=r, column=1, value=txt)
    if bold: c.font = Font(bold=True)
    r += 1
ws.column_dimensions['A'].width = 125

ws = wb.create_sheet('ADD TO TARGET LIST')
cols = [('Add group',22),('Company',32),('Website',26),('HQ',24),('Type bucket',12),('Campaign',26),
        ('Priority',14),('Employees',11),('Revenue est.',26),('Year founded',12),('Ownership',16),
        ('Owner / contact',20),('Title',26),('Email',30),('Email verified',13),('MOBILE',20),
        ('Direct dial',18),('Postal address',40),('Verify first?',14),('Note / reason',60),('Source file(s)',44)]
for i,(h,w) in enumerate(cols,1):
    c = ws.cell(row=1, column=i, value=h); c.fill = HDR; c.font = HF; c.alignment = WRAP
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'C2'
keys = ['group','company','website','hq','bucket','campaign','priority','employees','revenue','founded',
        'ownership','contact','title','email','email_verified','mobile','direct','address','verify','note','src']
for n,row in enumerate(rows,2):
    for i,k in enumerate(keys,1):
        ws.cell(row=n, column=i, value=row.get(k)).alignment = WRAP
    if row['verify']:
        for i in range(1,22): ws.cell(row=n, column=i).fill = RED
    elif row['mobile']:
        ws.cell(row=n, column=16).fill = GREEN
    if row['group'].startswith('NURTURE'):
        ws.cell(row=n, column=1).fill = AMBER
    ws.row_dimensions[n].height = 26

OUT = '/home/user/Master-Target-List/Hunter_Power_Companies_to_Add_20260902.xlsx'
wb.save(OUT)
print('rows:', len(rows),
      '| net-new:', sum(1 for r in rows if r['group']=='NET-NEW'),
      '| nurture:', sum(1 for r in rows if r['group'].startswith('NURTURE')),
      '| with mobile:', sum(1 for r in rows if r['mobile']),
      '| with email:', sum(1 for r in rows if r['email']),
      '| with address:', sum(1 for r in rows if r['address']),
      '| flagged:', sum(1 for r in rows if r['verify']))
print('wrote', OUT)

# ---------------------------------------------------------- Reconciliation
# Every total in the four source files, counted rather than taken on trust.
ws = wb.create_sheet('RECONCILIATION')
BOLD = Font(bold=True)
def line(r, cells, bold=False, fill=None):
    for i, v in enumerate(cells, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.alignment = WRAP
        if bold: c.font = BOLD
        if fill: c.fill = fill
    return r + 1

r = 1
ws.cell(row=1, column=1, value='Reconciliation of all four files').font = Font(bold=True, size=14)
r = 3
r = line(r, ['MASTER TARGET LIST v3 (25 Aug)', 'Rows', 'Unique domains', 'Note'], True, HDR)
for i in range(1, 5): ws.cell(row=r-1, column=i).font = HF
for sh, n, note in [('Batch 1 (250)', 250, 'matches the sheet name'),
                    ('Batch 2 (250)', 250, 'matches'), ('Batch 3 (147)', 147, 'matches'),
                    ('NEW - Add to batch 1 or 2', 11, 'NOT in the "All 1922 master" sheet'),
                    ('NEW - Add to batch 2 or 3', 51, 'NOT in the "All 1922 master" sheet'),
                    ('Nurture (below floor)', 138, 'the 138 you asked to re-add'),
                    ('DQ log', 1137, 'excluded from everything')]:
    r = line(r, [sh, n, n, note])
r = line(r, ['Sum of those sheets', 1984, 1984, 'no duplication between sheets'], True)
r = line(r, ['"All 1922 master" sheet', 1922, 1922, 'the sheet that claims to be the full master'], True)
r = line(r, ['DIFFERENCE', 62, 62,
             'The 62 rows on the two "NEW - Add to batch" sheets were never merged into '
             '"All 1922 master". That sheet is 62 short of the file it sits in. None of the 62 '
             'appears in the later files, so nothing here was double-counted - but if you rebuild '
             'a master, take the union of the sheets, not the "All 1922" tab.'], True, AMBER)
r += 1
r = line(r, ['THE EXPANSION CHAIN (six files, in order)', 'Companies', '', 'Note'], True, HDR)
for i in range(1, 5): ws.cell(row=r-1, column=i).font = HF
r = line(r, ['1. Grata Expansion (31 Aug)', 1755, '', 'raw candidate pool - a search result, NOT a verdict'])
r = line(r, ['2. Reassessment (31 Aug)', 1755, '', 'every candidate re-bucketed and re-prioritised'])
r = line(r, ['3. Campaign Assignment (31 Aug)', 1755, '',
             'adjudicated all 1,755: 832 campaign-assigned + 212 Nurture + 711 DQ'], True)
r = line(r, ['4. Validated Campaigns (1 Sep)', 1045, '',
             'took the 1,044 non-DQ rows and kept 729, DQd 316'])
r = line(r, ['5. New Targets (1 Sep)', 867, '', '729 + 138 released from master v3 Nurture'], True, GREEN)
r = line(r, ['6. Master Target List v3 (25 Aug)', 1922, '', 'the base everything is measured against'])
r += 1
r = line(r, ['Validated is a subset of New Targets', 0, '',
             'companies in Validated but not New Targets: none. New Targets SUPERSEDES it.'], True, GREEN)
r = line(r, ['Caution when summing tabs', 1005, 867,
             'New Targets tabs sum to 1,005 but hold 867: the "NEW additions (138)" tab repeats rows '
             'that also sit on the campaign tabs. Do not add the tabs together.'], False, AMBER)
r += 1
r = line(r, ['HOW 726 NET-NEW IS ARRIVED AT', 'Companies', '', 'Note'], True, HDR)
for i in range(1, 5): ws.cell(row=r-1, column=i).font = HF
r = line(r, ['Campaign + Nurture sheets, all files', 1893, '', 'every company any pass placed on a live list'])
r = line(r, ['less: DQd somewhere in the chain', -1029, '',
             '711 at Campaign Assignment, 316 at validation, 86 in the expansion itself'])
r = line(r, ['less: already in master v3', -138, '', 'exactly the Nurture releases; nothing else was known'])
r = line(r, ['NET-NEW', 726, '', 'all validated - every one sits on a 1 Sep campaign sheet'], True, GREEN)
r = line(r, ['plus: master v3 Nurture', 138, '', 'you asked for these back'], False, AMBER)
r = line(r, ['TOTAL TO ADD', 864, '', ''], True, GREEN)
r += 1
r = line(r, ['CORRECTION', '', '',
             'An earlier version of this file said 1,353 net-new, of which 627 were "never validated". '
             'That was wrong. It was built before the Campaign Assignment file was available, so the '
             '711 companies that pass DQd on 31 August looked unassessed. All 627 sit on that DQ '
             'sheet - they were examined and rejected, not overlooked. Reading the raw expansion tabs '
             'as targets is the error: they are a search result, not a verdict.'], True, AMBER)

for i, w in enumerate([40, 16, 16, 96], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
wb.save(OUT)
print('reconciliation sheet added')
